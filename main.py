import json
import multiprocessing
import os
import time
import re
import openpyxl
from datetime import datetime
import requests
from selenium.webdriver.chrome.service import Service
from selenium import webdriver
from bs4 import BeautifulSoup as bs
from multiprocessing import Process
from seleniumwire import webdriver as webdriver_wire
from fake_useragent import UserAgent

count_errors = 0


def SaveOZON(datenow, timenow, PRODUCT, VENDOR_CODE, FIND_STRING, position, page, price, comments, rating):
    ozon = openpyxl.reader.excel.load_workbook(filename="data/Результат.xlsx")
    ozon.active = 3
    ozon_requests = ozon.active
    maxrows = ozon_requests.max_row + 1
    if position != "9999":
        ozon_requests['A' + str(maxrows)] = datenow
        ozon_requests['B' + str(maxrows)] = timenow
        ozon_requests['C' + str(maxrows)] = PRODUCT
        ozon_requests['D' + str(maxrows)] = VENDOR_CODE
        ozon_requests['E' + str(maxrows)] = FIND_STRING
        ozon_requests['F' + str(maxrows)] = position
        ozon_requests['G' + str(maxrows)] = page
        ozon_requests['H' + str(maxrows)] = price
        ozon_requests['I' + str(maxrows)] = comments
        ozon_requests['J' + str(maxrows)] = rating
        print("Обьект записан")
        ozon.save("data/Результат.xlsx")
    else:
        ozon_requests['A' + str(maxrows)] = datenow
        ozon_requests['B' + str(maxrows)] = timenow
        ozon_requests['C' + str(maxrows)] = PRODUCT
        ozon_requests['D' + str(maxrows)] = VENDOR_CODE
        ozon_requests['E' + str(maxrows)] = FIND_STRING
        ozon_requests['F' + str(maxrows)] = position
        print("Обьект записан")
        ozon.save("data/Результат.xlsx")


def LoadOzon():
    wb = openpyxl.reader.excel.load_workbook(filename="data/Результат.xlsx")
    wb.active = 1
    wb_requests = wb.active
    maxrows = wb_requests.max_row
    for i in range(2, maxrows + 1):
        PRODUCT = wb_requests['A' + str(i)].value
        VENDOR_CODE = wb_requests['B' + str(i)].value
        FIND_STRING = wb_requests['C' + str(i)].value
        print(PRODUCT, VENDOR_CODE, FIND_STRING)

        times = [1, 2, 3]
        found_event = multiprocessing.Event()

        pool = [Process(target=OzonCheck, args=(FIND_STRING, VENDOR_CODE, PRODUCT, 0, number, 0, found_event, number))
                for number in times]

        for p in pool:
            p.start()

        found_event.wait()
        print('{} | terminating processes'.format(datetime.now()))
        for p in pool:
            p.terminate()
        for p in pool:
            p.join()
        print('{} | all processes joined'.format(datetime.now()))


def OzonCheck(FIND_STRING, VENDOR_CODE, PRODUCT, count=0, page_now=1, pagination=0, found_event="", check=0):
    url = f"https://www.ozon.ru/search/?from_global=true&page={page_now}&text={FIND_STRING}"
    print(f"Текущая страница: №{page_now}")
    options = webdriver.ChromeOptions()
    ua = UserAgent()
    options.add_argument(
        f"user-agent={ua.random}")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-gpu")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1920x1080")
    options.add_argument("start-maximised")
    service = Service(f"{os.getcwd()}\\chromedriver\\chromedriver.exe")
    driver = webdriver.Chrome(
        service=service,
        options=options
    )
    driver.implicitly_wait(2)
    try:
        driver.get(url=url)
        time.sleep(2)
    except Exception:
        print("Ошибка")
    soup = bs(driver.page_source, "lxml")
    if not soup.find('div', attrs={'data-widget': 'searchResultsError'}):
        cards_class_name_first = \
            soup.find('div', attrs={'data-widget': 'searchResultsV2'}).find(class_="ui-p6").parent.get("class")[0]
        cards_class_name_second = soup.find(class_=cards_class_name_first).parent.get("class")[0]
        cards = soup.find_all(class_=cards_class_name_second)
        for card in cards:
            count += 1
            s = bs(str(card), "lxml")
            try:
                href = "https://www.ozon.ru" + s.find("a").get("href")
            except  Exception:
                print("Ошибка")
                href = ""
                driver.quit()
            if str(VENDOR_CODE) in href:
                position = ((page_now - 1) * 36) + count
                price = s.find(class_="ui-p6").find(class_="ui-p9 ui-q1 ui-q4").text.replace(" ", "").replace(" ",
                                                                                                              "").replace(
                    "₽",
                    "").replace(" ", "")

                driver.get(url=href)
                time.sleep(2)
                s = driver.execute_script("return document.body.scrollHeight")
                s = int(s) / 2
                print(href)
                print(s)
                driver.execute_script(f"window.scrollTo(0, {s})")
                time.sleep(2)
                soupp = bs(driver.page_source, "lxml")
                try:
                    comments = re.findall('\d+', (
                        soupp.find('div', attrs={'data-widget': 'webReviewProductScore'}).find(class_="ui-e7").text.replace(" ", "").replace(" ", "")))[0]
                except Exception:
                    comments = "0"
                try:
                    rating = soupp.find('div', attrs={'data-widget': 'paginator'}).find('div', attrs={
                        'data-widget': 'webReviewProductScore'}).find("span").text.replace(".",
                                                                                           ",").split(
                        " / ", 1)[0]
                except Exception:
                    rating = "0"
                datenow = datetime.now().strftime("%d.%m.%Y")
                timenow = datetime.now().strftime("%H:%M:%S")
                print(position, page_now, price, comments, rating, datenow, timenow)
                SaveOZON(datenow, timenow, PRODUCT, VENDOR_CODE, FIND_STRING, position, page_now, price, comments,
                         rating)
                driver.quit()
                found_event.set()
                return
        driver.quit()
        if check == 1:
            if page_now == 20:
                page_now = 0
        page_now = page_now + 1
        OzonCheck(FIND_STRING, VENDOR_CODE, PRODUCT, count=0, page_now=page_now, pagination=pagination,
                  found_event=found_event, check=check)
    else:
        print("По этому запросу продукт не найден")
        driver.quit()
        datenow = datetime.now().strftime("%d.%m.%Y")
        timenow = datetime.now().strftime("%H:%M:%S")
        SaveOZON(datenow, timenow, PRODUCT, VENDOR_CODE, FIND_STRING, "9999", "", "", "", "",
                 "")


def SaveWB(datenow, timenow, PRODUCT, VENDOR_CODE, FIND_STRING, position, page, price, comments, rating, purchases):
    global count_errors
    count_errors = 0
    wb = openpyxl.reader.excel.load_workbook(filename="data/Результат.xlsx")
    wb.active = 2
    wb_requests = wb.active
    maxrows = wb_requests.max_row + 1
    if position != "9999":
        wb_requests['A' + str(maxrows)] = datenow
        wb_requests['B' + str(maxrows)] = timenow
        wb_requests['C' + str(maxrows)] = PRODUCT
        wb_requests['D' + str(maxrows)] = VENDOR_CODE
        wb_requests['E' + str(maxrows)] = FIND_STRING
        wb_requests['F' + str(maxrows)] = position
        wb_requests['G' + str(maxrows)] = page
        wb_requests['H' + str(maxrows)] = price
        wb_requests['I' + str(maxrows)] = comments
        wb_requests['J' + str(maxrows)] = rating
        wb_requests['K' + str(maxrows)] = purchases
        print("Обьект записан")
        wb.save("data/Результат.xlsx")
    else:
        wb_requests['A' + str(maxrows)] = datenow
        wb_requests['B' + str(maxrows)] = timenow
        wb_requests['C' + str(maxrows)] = PRODUCT
        wb_requests['D' + str(maxrows)] = VENDOR_CODE
        wb_requests['E' + str(maxrows)] = FIND_STRING
        wb_requests['F' + str(maxrows)] = position
        print("Обьект записан")
        wb.save("data/Результат.xlsx")


def LoadWB():
    wb = openpyxl.reader.excel.load_workbook(filename="data/Результат.xlsx")
    wb.active = 0
    wb_requests = wb.active
    maxrows = wb_requests.max_row
    for i in range(2, maxrows + 1):
        PRODUCT = wb_requests['A' + str(i)].value
        VENDOR_CODE = wb_requests['B' + str(i)].value
        FIND_STRING = wb_requests['C' + str(i)].value
        print(PRODUCT, VENDOR_CODE, FIND_STRING)
        FindWB(FIND_STRING, VENDOR_CODE, PRODUCT)


def main():
    print("Загружаем WildBerries")
    LoadWB()
    print("Загружаем Ozon")
    LoadOzon()
    print("БОТ ЗАКОНЧИЛ РАБОТУ!")


def FindWB(FIND_STRING, VENDOR_CODE, PRODUCT, count=0):
    global count_errors
    options = webdriver_wire.ChromeOptions()
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.159 Safari/537.36 OPR/78.0.4093.184")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-gpu")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1920x1080")
    options.add_argument("start-maximised")
    driver = webdriver_wire.Chrome(executable_path=f"{os.getcwd()}\\chromedriver\\chromedriver.exe", options=options)
    json_url = ""
    driver.get(
        f'https://www.wildberries.ru/catalog/0/search.aspx?page=1&search={FIND_STRING}')
    driver.implicitly_wait(2)
    time.sleep(2)
    for request in driver.requests:
        if request.response:
            if "https://wbxcatalog-ru.wildberries.ru/" in request.url:
                json_url = request.url
                break
    driver.quit()
    running = True
    page = 1
    while running:
        try:
            sep = 'page='
            json_url = json_url.split(sep, 1)[0]
            json_url += f"page={page}"
            req = requests.get(json_url)
            json_string = req.text
            data = json.loads(json_string)
            for i in data['data']['products']:
                count += 1
                s = str(i.get("id"))
                if str(s) == str(VENDOR_CODE):
                    running = False
                    print(f"Страница: {page}")
                    print(count)
                    url = f"https://www.wildberries.ru/catalog/{VENDOR_CODE}/detail.aspx?targetUrl=SP"
                    LoadInfoWB(FIND_STRING=FIND_STRING, VENDOR_CODE=VENDOR_CODE, PRODUCT=PRODUCT, url=url, page=page,
                               count=count)
            page += 1
        except Exception:
            if running == True:
                count_errors += 1
                print(count_errors)
                if count_errors < 4:
                    running = False
                    print("Еще одна проверка")
                    FindWB(FIND_STRING, VENDOR_CODE, PRODUCT)
                else:
                    running = False
                    # (f"Последняя страница: {page - 1}")
                    print("Все страницы прошел")
                    datenow = datetime.now().strftime("%d.%m.%Y")
                    timenow = datetime.now().strftime("%H:%M:%S")
                    SaveWB(datenow, timenow, PRODUCT, VENDOR_CODE, FIND_STRING, "9999", "", "", "", "",
                           "")


def LoadInfoWB(FIND_STRING, VENDOR_CODE, PRODUCT, url, page, count):
    options = webdriver.ChromeOptions()
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.159 Safari/537.36 OPR/78.0.4093.184")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1920x1080")
    options.add_argument("start-maximised")
    service = Service(f"{os.getcwd()}\\chromedriver\\chromedriver.exe")
    driver = webdriver.Chrome(
        service=service,
        options=options
    )

    driver.implicitly_wait(1)
    try:
        driver.get(url=url)
        time.sleep(2)
    except Exception:
        print("Ошибка")
    s = driver.execute_script("return document.body.scrollHeight")
    driver.execute_script("window.scrollTo(0, {0})".format(s))
    time.sleep(2)
    soupp = bs(driver.page_source, "lxml")
    price = soupp.find(class_="price-block__final-price").text.replace(" ", "").replace(" ",
                                                                                        "").replace("₽",
                                                                                                    "")
    try:
        comments = re.findall('\d+', (
            soupp.find(class_="same-part-kt__count-review").text))[0]
    except Exception:
        comments = "0"
    try:
        rating = soupp.find(class_="user-scores__score").text.replace(".", ",")
    except Exception:
        rating = "0"
    try:
        purchases = re.findall('\d+', (
            soupp.find(class_="same-part-kt__order-quantity j-orders-count-wrapper").text.replace(" ", "").replace(
                " ",
                "")))[0]
    except Exception:
        purchases = "0"
    datenow = datetime.now().strftime("%d.%m.%Y")
    timenow = datetime.now().strftime("%H:%M:%S")
    print(url, count, page, price, comments, rating, purchases, datenow, timenow)
    SaveWB(datenow, timenow, PRODUCT, VENDOR_CODE, FIND_STRING, count, page, price, comments, rating,
           purchases)


if __name__ == '__main__':
    main()
